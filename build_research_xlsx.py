import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "מקורות מחקר"
ws.sheet_view.rightToLeft = True

ws.append(["נכנס?", "נושא", "רמה", "כותרת", "קישור", "למה רלוונטי"])
hdr_fill = PatternFill("solid", fgColor="1D9E75")
hdr_font = Font(bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")

rows = [
("ריצות קלות / Zone 2","אקדמי","What Is Zone 2 Training? Experts' Viewpoint (IJSPP 2025)","https://journals.humankinetics.com/view/journals/ijspp/20/11/article-p1614.xml","קונצנזוס 14 מדענים על הגדרת Zone 2"),
("ריצות קלות / Zone 2","אקדמי","Zone 2 Individual Variability (Meixner, TSM 2025)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC11986187/","שונות מעל 20% בדופק - צורך בהתאמה אישית"),
("ריצות קלות / Zone 2","אקדמי","Training Intensity Distribution elite (Frontiers 2015)","https://www.frontiersin.org/journals/physiology/articles/10.3389/fphys.2015.00295/full","סקירת 80/20 על פני ענפי סיבולת"),
("ריצות קלות / Zone 2","אקדמי","Polarized > threshold/HIIT/volume (Stoggl PMC 2014)","https://pmc.ncbi.nlm.nih.gov/articles/PMC3912323/","ניסוי: polarized = עליית VO2peak מקסימלית"),
("ריצות קלות / Zone 2","אקדמי","Quantifying TID optimal (Seiler 2006)","https://pubmed.ncbi.nlm.nih.gov/16430681/","המאמר המכונן של מודל 80/20"),
("ריצות קלות / Zone 2","מאומת","FastTalk Labs - Polarized Training w/ Seiler","https://www.fasttalklabs.com/pathways/polarized-training/","מקור ראשון מ-Seiler עצמו"),
("ריצות קלות / Zone 2","מאומת","TrainingPeaks - Does Polarized Training Work?","https://www.trainingpeaks.com/blog/does-polarized-training-really-work/","פלטפורמת אימון מובילה"),
("ריצות קלות / Zone 2","מאומת","UESCA - Is Zone 2 the Endurance Panacea?","https://uesca.com/is-zone-2-really-the-endurance-panacea/","ארגון הסמכת מאמנים, מבט ביקורתי"),
("ריצות קלות / Zone 2","מאומת","HPRC - What's Zone 2 Training","https://www.hprc-online.org/physical-fitness/training-performance/whats-zone-2-training-and-why-does-it-matter","גוף בריאות צבא ארה'ב"),
("ריצות קלות / Zone 2","מאומת","Mountain Tactical - Zone 1 priority review","https://mtntactical.com/knowledge/research-review-endurance-programming-that-prioritizes-zone-1-produces-the-greatest-performance-gains-in-elite-endurance-athletes/","סקירת מחקר על עדיפות עצימות נמוכה"),
("ריצות קלות / Zone 2","אמין-למראה","TRX - Zone 2 Science-Backed Plan","https://www.trxtraining.com/blogs/news/zone-2-training-science-backed-aerobic-capacity-plan","מדריך מעשי"),
("ריצות קלות / Zone 2","אמין-למראה","Mountain Tactical - Mini Study 3h/week Zone 2","https://mtntactical.com/research/mini-study-3-hours-of-weekly-zone-2-running-improves-aerobic-base/","ניסוי שדה קטן"),

("ריצות כבדות / אינטרוולים","אקדמי","Aerobic HIIT > sprint for VO2max (Hov 2023)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC10099854/","אינטרוולים אירוביים ארוכים עדיפים ל-VO2max"),
("ריצות כבדות / אינטרוולים","אקדמי","Manipulating HIIT: VO2max/LT/3000m (2006)","https://pubmed.ncbi.nlm.nih.gov/16876479/","השפעת פרמטרי HIIT על ביצועים"),
("ריצות כבדות / אינטרוולים","אקדמי","HIIT vs Continuous VO2max meta-analysis","https://www.researchgate.net/publication/279980252_Effectiveness_of_High-Intensity_Interval_Training_HIT_and_Continuous_Endurance_Training_for_VO2max_Improvements_A_Systematic_Review_and_Meta-Analysis_of_Controlled_Trials","מטא-אנליזה השוואתית"),
("ריצות כבדות / אינטרוולים","אקדמי","Lactate Threshold Training review (MJSSM 2024)","https://mjssm.me/clanci/MJSSM_March_2024_Vijay.pdf","אימון סף לריצות ארוכות"),
("ריצות כבדות / אינטרוולים","אקדמי","Fixed Lactate Thresholds responses (PMC 2023)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC10611166/","תגובות אישיות באחוז VO2max בסף"),
("ריצות כבדות / אינטרוולים","מאומת","RunnersConnect - Lactate Threshold vs VO2max","https://runnersconnect.net/lactate-vo2max/","אתר אימון מבוסס-מחקר"),
("ריצות כבדות / אינטרוולים","מאומת","RunBikeCalc - Interval Training Guide 2026","https://runbikecalc.com/blog/interval-training-complete-guide-2026","מדריך HIIT מקיף"),
("ריצות כבדות / אינטרוולים","מאומת","Arya Athletics - Lactate Threshold Guide","https://www.arya-athletics.com/post/lactate-threshold-training-for-distance-runners-a-science-based-guide","מאמן אונליין מבוסס-מדע"),
("ריצות כבדות / אינטרוולים","אמין-למראה","Shuichi Running - Lactate Threshold Guide","https://shuichi-running.com/en/lactate-threshold-training/","מדריך מפורט"),
("ריצות כבדות / אינטרוולים","אמין-למראה","Vital Performance - Mastering the Tempo Run","https://vitalperformancecare.com/mastering-the-tempo-run-the-key-to-faster-stronger-running/","מדריך tempo"),

("תחרויות / Taper","אקדמי","Tapering meta-analysis (PLOS One 2023)","https://pmc.ncbi.nlm.nih.gov/articles/PMC10171681/","המקור המרכזי: 8-14 ימים, נפח מינוס 41-60%"),
("תחרויות / Taper","אקדמי","Tapering Elite Endurance Runners review (2018)","https://www.researchgate.net/publication/322931606_The_Effects_of_Tapering_on_Performance_in_Elite_Endurance_Runners_A_Systematic_Review","סקירה שיטתית על רצים מובילים"),
("תחרויות / Taper","אקדמי","Nutrition Strategies for the Marathon (Burke)","https://www.researchgate.net/publication/6363538_Nutrition_Strategies_for_the_Marathon_Fuel_for_Training_and_Racing","carbo-loading 10-12 גרם לקג"),
("תחרויות / Taper","אקדמי","Controlling nutrition & force in long race (Frontiers 2023)","https://www.frontiersin.org/journals/nutrition/articles/10.3389/fnut.2023.1096194/full","מודל אופטימיזציה של דלק וקצב"),
("תחרויות / Taper","אקדמי","Pacing marathon vs half-marathon (PMC 2024)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC10771621/","negative/even splits מול positive"),
("תחרויות / Taper","מאומת","Science in Sport - In-Race Fuelling Marathon","https://www.scienceinsport.com/sports-nutrition/mastering-in-race-fuelling-for-marathon-success/","מותג תזונת ספורט מוביל"),
("תחרויות / Taper","מאומת","RunnersConnect - Marathon Bonking/Glycogen","https://runnersconnect.net/glycogen-depletion-marathon/","מניעת קריסה אנרגטית"),
("תחרויות / Taper","אמין-למראה","Pneux - Marathon Fueling Strategy 2026","https://pneux.io/marathon-fueling-strategy","אסטרטגיית דלק מעשית"),
("תחרויות / Taper","אמין-למראה","Zolt - Marathon Nutrition Complete Guide","https://www.zolthealth.com/blog/marathon-training-nutrition-complete-fueling-guide","מדריך תזונה מקיף"),
("תחרויות / Taper","אמין-למראה","Sport-Calculator - Race Planner","https://sport-calculator.com/tools/running-race-planner","כלי תכנון מרוץ"),

("התאוששות / Recovery","אקדמי","Sleep & Nightly Recovery wearables (PMC 2025)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC11768492/","קשר שינה-עומס-הסתגלות"),
("התאוששות / Recovery","אקדמי","Overtraining Syndrome 3 athletes (PMC 2024)","https://pmc.ncbi.nlm.nih.gov/articles/PMC12258013/","case studies של OTS והחלמה"),
("התאוששות / Recovery","אקדמי","HRV via Mobile review (MDPI Sensors)","https://www.mdpi.com/1424-8220/26/1/3","איך ומתי למדוד HRV"),
("התאוששות / Recovery","אקדמי","Predicting daily recovery ML (Springer 2024)","https://link.springer.com/article/10.1007/s00421-024-05530-2","ניבוי החלמה יומית"),
("התאוששות / Recovery","אקדמי","One-week deload adaptations (PeerJ)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC10809978/","דה-לואוד שומר היפרטרופיה"),
("התאוששות / Recovery","מאומת","Built With Science - Deload Weeks Explained","https://builtwithscience.com/fitness-tips/deload-week/","מבוסס-מחקר, מוניטין גדול"),
("התאוששות / Recovery","מאומת","Mundo Entrenamiento - Deload Practical Guide","https://mundoentrenamiento.com/en/deload-week/","מדריך מעשי"),
("התאוששות / Recovery","אמין-למראה","RunDida - Supercompensation Theory","https://rundida.com/guide/supercompensation-recovery/","תזמון התאוששות"),
("התאוששות / Recovery","אמין-למראה","FitnessRec - Supercompensation Timing","https://fitnessrec.com/articles/supercompensation-theory-for-athletes-master-recovery-timing-for-maximum-gains","תזמון לרווח מקסימלי"),

("היברידי / כוח+סיבולת","אקדמי","Concurrent training sequence meta (Frontiers 2023)","https://www.frontiersin.org/journals/physiology/articles/10.3389/fphys.2023.1072679/full","כוח לפני סיבולת משמר כוח"),
("היברידי / כוח+סיבולת","אקדמי","Concurrent: sex & training status (PMC 2024)","https://pmc.ncbi.nlm.nih.gov/articles/PMC10933151/","interference קטן בגברים"),
("היברידי / כוח+סיבולת","אקדמי","Strength & runners' economy meta (PMC 2024)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC11052887/","כוח ופליומטרי משפר running economy"),
("היברידי / כוח+סיבולת","אקדמי","Heavy Resistance vs Plyometric economy (Springer 2022)","https://link.springer.com/article/10.1186/s40798-022-00511-1","כוח כבד בקצב נמוך, פליו בגבוה"),
("היברידי / כוח+סיבולת","אקדמי","Strength on distance running determinants (PMC 2018)","https://www.ncbi.nlm.nih.gov/pmc/articles/PMC5889786/","סקירה על מנגנונים פיזיולוגיים"),
("היברידי / כוח+סיבולת","מאומת","Stronger by Science - Interference Effect","https://www.strongerbyscience.com/research-spotlight-interference-effect/","מקור מבוסס-מחקר מוביל"),
("היברידי / כוח+סיבולת","מאומת","Barbell Medicine - Concurrent Training","https://www.barbellmedicine.com/blog/concurrent-training-and-the-interference-effect/","רופאים ומאמנים, סמכות גבוהה"),
("היברידי / כוח+סיבולת","מאומת","Polar - The Hybrid Athlete","https://www.polar.com/en/journal/hybrid-athlete","מותג שעוני ספורט"),
("היברידי / כוח+סיבולת","מאומת","Red Bull - Hybrid Training Guide","https://www.redbull.com/us-en/hybrid-training-guide","תוכן ספורט מקצועי"),
("היברידי / כוח+סיבולת","מאומת","Gymshark - What Is Hybrid Training","https://www.gymshark.com/blog/article/what-is-hybrid-training-complete-guide","מדריך מקיף"),
("היברידי / כוח+סיבולת","אמין-למראה","Tailored Coaching - Hybrid Step-by-Step","https://tailoredcoachingmethod.com/hybrid-athlete-training-guide/","עקרון High-Low קשה-קל"),
("היברידי / כוח+סיבולת","אמין-למראה","AthleteData - Hybrid Athlete Guide","https://www.athletedata.health/guides/hybrid-athlete-training","מדריך מפורט"),
("היברידי / כוח+סיבולת","אמין-למראה","TRX - Hybrid Athlete 12-Week Program","https://www.trxtraining.com/blogs/news/hybrid-athlete-training-program","תוכנית 12 שבועות"),
("היברידי / כוח+סיבולת","אמין-למראה","Nick Bare - Training for Hybrid Athletes","https://www.nickbare.com/training/","אתלט היברידי ידוע"),
("היברידי / כוח+סיבולת","אמין-למראה","Get Sprints - Running & Lifting Program","https://getsprints.com/blogs/get-inspired/the-ultimate-running-and-lifting-program-for-maximum-hybrid-athlete-performance","תוכנית משולבת"),
]

level_fill = {
 "אקדמי": PatternFill("solid", fgColor="D6EAF8"),
 "מאומת": PatternFill("solid", fgColor="D5F5E3"),
 "אמין-למראה": PatternFill("solid", fgColor="FCF3CF"),
}

for r in rows:
    ws.append(["☐", r[0], r[1], r[2], r[3], r[4]])
    i = ws.max_row
    ws.cell(row=i, column=4).hyperlink = r[3]
    ws.cell(row=i, column=4).font = Font(color="1155CC", underline="single")
    ws.cell(row=i, column=5).hyperlink = r[3]
    ws.cell(row=i, column=5).font = Font(color="1155CC", underline="single", size=9)
    ws.cell(row=i, column=3).fill = level_fill.get(r[1], PatternFill())
    for col in range(1, 7):
        ws.cell(row=i, column=col).border = border
        ws.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="center")

for i, w in enumerate([8, 24, 14, 50, 40, 38], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 24

wb.save("TRAINING_RESEARCH.xlsx")
print("Excel saved:", len(rows), "sources")
